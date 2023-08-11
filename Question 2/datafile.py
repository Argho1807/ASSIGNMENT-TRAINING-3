### QUESTION-2 ###

import math
import numpy as np
import pandas as pd
import time
from datetime import datetime
import random

import gurobipy as grb
import openpyxl
from openpyxl.styles import Alignment
import os

# MODEL FORMULATION USING GUROBI SOLVER

### SETS AND INDICES ###

num_items = 15
num_vendors = 5
num_SD = 3
num_QD = 3
M = 9999999999999999999999999999999999999999 # big-M
E = 1/M # a very small value

I = [f'Item {i+1}' for i in range(num_items)]
V = [f'Vendor {v+1}' for v in range(num_vendors)]
SD = ['<S1', '>=S1, <S2', '>=S2']
QD = ['<Q1', '>=Q1, <Q2', '>=Q2']

### PARAMETERS ###

seed_val = 20
np.random.seed(seed_val)

Demand = [np.random.uniform(200, 400) for i in range(num_items)]
Quantity = [[np.random.uniform(0, 100) for i in range(num_items)] for v in range(num_vendors)]
Price = [[np.random.uniform(6, 10) for i in range(num_items)] for v in range(num_vendors)]

BidPrice = [sum(Price[v][i]*Quantity[v][i] for i in range(num_items)) for v in range(num_vendors)]
TotalQ = [sum(Quantity[v][i] for i in range(num_items)) for v in range(num_vendors)]

S1 = 1/3*sum(BidPrice)/num_vendors
S2 = 2/3*sum(BidPrice)/num_vendors
x1, x2 = 0.1, 0.15

Q1 = 1/3*sum(TotalQ)/num_vendors
Q2 = 2/3*sum(TotalQ)/num_vendors
y1, y2 = 0.05, 0.1

PenaltyCost = [sum(Price[v][i] for v in range(num_vendors)) for i in range(num_items)] # a large number

# Saving input data to a new worksheet
inputworkbook = openpyxl.Workbook()
sheet = inputworkbook.active

for i in range(num_items):
    sheet.cell(row = 1, column = 1, value = 'ITEM')
    sheet.cell(row = 2, column = 1, value = 'DEMAND')
    sheet.cell(row = 1, column = i+2, value = I[i])
    sheet.cell(row = 2, column = i+2, value = Demand[i])

    for v in range(num_vendors):
        sheet.cell(row = 4, column = 1, value = 'QUANTITY')
        sheet.cell(row = i+5, column = 1, value = I[i])
        sheet.cell(row = 4, column = v+2, value = V[v])
        sheet.cell(row = i+5, column = v+2, value = Quantity[v][i])
        
        sheet.cell(row = 4+num_items+2, column = 1, value = 'PRICE')
        sheet.cell(row = i+5+num_items+2, column = 1, value = I[i])
        sheet.cell(row = 4+num_items+2, column = v+2, value = V[v])
        sheet.cell(row = i+5+num_items+2, column = v+2, value = Price[v][i])

inputworkbook.save('input.xlsx')
os.startfile('input.xlsx') # opens the output excel file